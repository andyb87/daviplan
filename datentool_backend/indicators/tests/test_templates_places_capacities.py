import os
from io import BytesIO
from openpyxl.reader.excel import load_workbook
import pandas as pd
import unittest

from django.urls import reverse
from test_plus import APITestCase
from datentool_backend.api_test import LoginTestCase
from datentool_backend.infrastructure.factories import (InfrastructureFactory,
                                                        ServiceFactory,
                                                        Place,
                                                        CapacityFactory,
                                                        PlaceFieldFactory,
                                                        FieldTypeFactory,
                                                        PlaceFactory,
                                                        )

from datentool_backend.area.models import FieldTypes
from datentool_backend.area.factories import FClassFactory


class InfrastructureTemplateTest(LoginTestCase, APITestCase):

    testdata_folder = 'testdata'

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.profile.can_edit_basedata = True
        cls.profile.save()
        cls.infra = InfrastructureFactory()
        cls.service1 = ServiceFactory(name='Schulen',
                                      has_capacity=False,
                                      infrastructure=cls.infra)
        cls.service2 = ServiceFactory(name='Schulplätze',
                                      has_capacity=True,
                                      infrastructure=cls.infra)

        cls.str_field = PlaceFieldFactory(infrastructure=cls.infra,
                                          field_type__ftype=FieldTypes.STRING,
                                          name='Bezeichnung',
                                          is_label=True)
        cls.hnr_field = PlaceFieldFactory(infrastructure=cls.infra,
                                          name='Hausnummer',
                                          field_type__ftype=FieldTypes.STRING)
        cls.num_field = PlaceFieldFactory(infrastructure=cls.infra,
                                          name='Gewichtung',
                                          field_type__ftype=FieldTypes.NUMBER)

        cl_ft1 = FieldTypeFactory(ftype=FieldTypes.CLASSIFICATION)
        cl_ft2 = FieldTypeFactory(ftype=FieldTypes.CLASSIFICATION)
        cls.cla_field = PlaceFieldFactory(infrastructure=cls.infra,
                                          name='EinsZweiDrei',
                                          field_type=cl_ft1)

        cv11 = FClassFactory(ftype=cls.cla_field.field_type, value='Eins', order=1)
        cv12 = FClassFactory(ftype=cls.cla_field.field_type, value='Zwei', order=2)
        cv13 = FClassFactory(ftype=cls.cla_field.field_type, value='Drei', order=3)

        cls.cla_field2 = PlaceFieldFactory(infrastructure=cls.infra,
                                           name='AABB',
                                           field_type=cl_ft2)

        cv21 = FClassFactory(ftype=cls.cla_field2.field_type, value='AA', order=1)
        cv22 = FClassFactory(ftype=cls.cla_field2.field_type, value='BB', order=2)

        place1 = PlaceFactory(infrastructure=cls.infra, attributes={
            cls.str_field.name: 'Place1',
            cls.num_field.name: 1234,
            cls.hnr_field.name: '44b',
            cls.cla_field.name: 'Zwei',
            cls.cla_field2.name: 'BB',
        })
        place2 = PlaceFactory(infrastructure=cls.infra, attributes={
            cls.str_field.name: 'Place2',
            cls.num_field.name: 567,
            cls.hnr_field.name: '33',
            cls.cla_field.name: 'Drei',
        })

        CapacityFactory(service=cls.service1, place=place1, capacity=1)
        CapacityFactory(service=cls.service1, place=place2, capacity=0)
        CapacityFactory(service=cls.service2, place=place1, capacity=55.5)
        CapacityFactory(service=cls.service2, place=place2, capacity=0)

    def test_create_infrastructure_template(self):
        url = reverse('infrastructures-create-template', kwargs={'pk':self.infra.pk})
        res = self.post(url)
        self.assert_http_200_ok(res)
        wb = load_workbook(BytesIO(res.content))
        self.assertSetEqual(set(wb.sheetnames),
                            {'Standorte und Kapazitäten', 'meta', 'Klassifizierungen'})

    @unittest.skip('Not implemented yet')
    def test_upload_place_template(self):
        """
        test bulk upload places and capacities
        """
        file_name_places = 'Standorte_und_Kapazitäten.xlsx'
        file_path_places = os.path.join(os.path.dirname(__file__),
                                        self.testdata_folder,
                                        file_name_places)
        file_content = open(file_path_places, 'rb')
        data = {
            'excel_file' : file_content,
        }

        url = reverse('infrastructures-upload-template', kwargs={'pk':self.infra.pk})
        res = self.client.post(url, data,
                               extra=dict(format='multipart/form-data'))
        self.assert_http_202_accepted(res, msg=res.content)
        # 4 stops should have been uploaded with the correct hst-nr and names
        df = pd.read_excel(file_path_places, skiprows=[1])

        actual = pd.DataFrame(Place.objects.values('id', 'name')).set_index('id')
        expected = df[['HstNr', 'HstName']]\
            .rename(columns={'HstNr': 'id','HstName': 'name',})\
            .set_index('id')
        pd.testing.assert_frame_equal(actual, expected)

        ## assert that without edit_basedata-permissions no upload is possible
        #self.profile.can_edit_basedata = False
        #self.profile.save()
        #res = self.client.post(url, data, pk=self.infra.pk,
        # extra=dict(format='multipart/form-data'))
        #self.assert_http_403_forbidden(res)

